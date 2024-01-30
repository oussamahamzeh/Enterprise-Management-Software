from django.contrib.auth.models import User
from cashier.views import line_split
from .models import Expense
from datetime import date
from .models import TransactionExportNumber
import os
import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from Inventory.models import Transaction, Item, Client
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment


def export_transactions(transactions):
    if transactions:
        print("Exporting Transactions!")

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(BASE_DIR, 'media/Templates/export_transactions.xlsx')
        save_path = os.path.join(BASE_DIR, 'media/Transactions Export')

        if not os.path.exists(save_path):
            os.makedirs(save_path)

        counter = TransactionExportNumber.get_next_export_number()

        # Apply center alignment to the cell
        alignment = Alignment(horizontal='center', vertical='center')

        font = Font(name='Arial', size=11, bold=True, italic=False)  # , color="0033CC")  # Customize font settings
        font02 = Font(name='Arial', size=12, bold=True, italic=True)  # , color="0033CC")  # Customize font settings

        # Load the template workbook
        template_workbook = openpyxl.load_workbook(template_path)
        worksheet = template_workbook[template_workbook.sheetnames[0]]

        # Get the current datetime
        current_datetime = datetime.now()

        # Format the datetime as a string
        formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        worksheet.cell(row=7, column=2, value=formatted_datetime)

        # Find the row where data starts in your template
        data_start_row = 11

        total_sum = 0
        added_rows = 1
        discount = 0
        rounded_discount = 0
        total_quantity = 0
        # Write the transactions to the worksheet
        for row_num, transaction in enumerate(transactions, data_start_row):
            added_rows = added_rows
            worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity)
            worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity).alignment = alignment
            total_quantity += int(transaction.quantity)
            if transaction.discount < 0:
                temp_value = abs(float(transaction.item.selling_price) - float(transaction.selling_price))
                worksheet.cell(row=row_num + added_rows, column=3, value=transaction.item.selling_price + temp_value)
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=transaction.item.selling_price + temp_value).alignment = alignment
            else:
                worksheet.cell(row=row_num + added_rows, column=3, value=transaction.item.selling_price)
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=transaction.item.selling_price).alignment = alignment
            lines = line_split(transaction.item.name)
            for idx, line in enumerate(lines):
                if idx != 0:
                    added_rows = added_rows + 1
                worksheet.cell(row=row_num + added_rows, column=1, value=line)

            total_sum = total_sum + (float(transaction.selling_price) * int(transaction.quantity))

            if transaction.discount > 0:
                discount = ((float(transaction.discount) / float(transaction.quantity)) / float(
                    transaction.item.selling_price)) * 100
                rounded_discount = round(discount, 2)
                added_rows = added_rows + 1
                worksheet.cell(row=row_num + added_rows, column=1, value="Discount " + str(rounded_discount) + " %")
                worksheet.cell(row=row_num + added_rows, column=1,
                               value="Discount " + str(rounded_discount) + " %").alignment = alignment
                worksheet.cell(row=row_num + added_rows, column=2, value="-")
                worksheet.cell(row=row_num + added_rows, column=2, value="-").alignment = alignment
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=float(transaction.item.selling_price) - float(transaction.selling_price))
                worksheet.cell(row=row_num + added_rows, column=3, value=float(transaction.item.selling_price) - float(
                    transaction.selling_price)).alignment = alignment

        total_sum = round(total_sum, 2)
        final_row = len(transactions) + 10 + added_rows
        worksheet.cell(row=final_row + 2, column=1, value="Sub-total")
        worksheet.cell(row=final_row + 2, column=1, value="Sub-total").font = font
        worksheet.cell(row=final_row + 2, column=2, value=total_sum * 0.89)
        worksheet.cell(row=final_row + 2, column=3, value="USD")
        worksheet.cell(row=final_row + 3, column=1, value="TVA")
        worksheet.cell(row=final_row + 3, column=1, value="TVA").font = font
        worksheet.cell(row=final_row + 3, column=2, value=total_sum * 0.11)
        worksheet.cell(row=final_row + 3, column=3, value="USD")
        worksheet.cell(row=final_row + 5, column=1, value="Total")
        worksheet.cell(row=final_row + 5, column=1, value="Total").font = font
        worksheet.cell(row=final_row + 5, column=2, value=total_sum)
        worksheet.cell(row=final_row + 5, column=3, value="USD")
        worksheet.cell(row=final_row + 6, column=1, value="Quantity")
        worksheet.cell(row=final_row + 6, column=1, value="Quantity").font = font
        worksheet.cell(row=final_row + 6, column=2, value=total_quantity)
        if total_quantity == 1:
            worksheet.cell(row=final_row + 6, column=3, value="Item")
        else:
            worksheet.cell(row=final_row + 6, column=3, value="Items")
        worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN")
        worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN").font = font02
        worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!")
        worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!").font = font02

        # Save the filled workbook
        file_name = f'Export_{counter}.xlsx'
        full_file_path = os.path.join(save_path, file_name)
        template_workbook.save(full_file_path)

        return HttpResponse(status=204)
    else:
        return HttpResponse(status=400)


def transaction_list(request):
    transactions = Transaction.objects.all().order_by('-trans_id')
    items = Item.objects.all()
    clients = Client.objects.all()
    users = User.objects.all()
    types = Transaction.objects.values_list('type', flat=True).distinct()

    if request.method == 'GET':
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        item = request.GET.get('item')
        client_name = request.GET.get('client')
        user = request.GET.get('user')
        type = request.GET.get('type')

        if from_date and to_date:
            to_date = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
            transactions = Transaction.objects.filter(time__range=(from_date, to_date))

        if item:
            transactions = transactions.filter(item=item)

        if client_name:
            client = get_object_or_404(Client, name=client_name)
            transactions = transactions.filter(client=client)

        if user:
            transactions = transactions.filter(user=user)

        if type:
            transactions = transactions.filter(type=type)

        # Check if the "export" parameter is in the request
        # Check if the "export" parameter is in the request
        export_type = request.GET.get('export')
        if export_type == 'excel':
            # Perform actions for Excel export (a)
            export_transactions(transactions)
        elif export_type == 'pdf':
            # Perform actions for PDF export (b)
            export_transactions(transactions)

    return render(request, 'transaction_list.html',
                  {'items': items, 'clients': clients, 'users': users, 'types': types, 'transactions': transactions})


def balance_sheet(request):
    items = Item.objects.all()
    if request.method == 'GET':
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        cash = 0
        profit = 0
        inventory_purchase_cost = 0
        shipping_cost = 0
        internal_expenses = 0
        net_profit = 0
        discounts = 0
        asset_value = 0
        transactions_count = 0
        quantity_sold = 0
        items_sold = {}

        for item in items:
            inventory_purchase_cost += (item.purchase_cost * int(item.quantity))
            asset_value += ((item.purchase_cost + item.shipping_cost) * int(item.quantity))

        if from_date and to_date:
            to_date = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
            transactions = Transaction.objects.filter(time__range=(from_date, to_date))
            expenses = Expense.objects.filter(time__range=(from_date, to_date))

            if transactions:
                for transaction in transactions:
                    if transaction.type == "transfer" or transaction.type == "wholesale":
                        cash += transaction.selling_price * transaction.quantity
                        transactions_count += 1
                        quantity_sold += transaction.quantity
                        shipping_cost += transaction.item.shipping_cost * int(transaction.quantity)
                        if transaction.item in items_sold:
                            items_sold[transaction.item] += transaction.quantity
                        else:
                            items_sold[transaction.item] = transaction.quantity

                    elif transaction.type == "return":
                        cash -= transaction.selling_price * transaction.quantity
                        transactions_count -= 1
                        quantity_sold -= transaction.quantity
                        if transaction.item in items_sold:
                            items_sold[transaction.item] -= transaction.quantity
                        else:
                            items_sold[transaction.item] = -transaction.quantity

                    profit += transaction.profit
                    discounts += transaction.discount

            if expenses:
                for expense in expenses:
                    internal_expenses += expense.amount

            net_profit = profit - internal_expenses

        context = {
            'cash': round(cash, 2),
            'profit': round(profit, 2),
            'inventory_purchase_cost': round(inventory_purchase_cost, 2),
            'shipping_cost': round(shipping_cost, 2),
            'internal_expenses': round(internal_expenses, 2),
            'net_profit': round(net_profit, 2),
            'discounts': round(discounts, 2),
            'asset_value': round(asset_value, 2),
            'transactions_count': round(transactions_count, 2),
            'quantity_sold': round(quantity_sold, 2),
            'items_sold': items_sold,
            'from_date': request.GET.get('from_date', ''),
            'to_date': request.GET.get('to_date', ''),
            'today': date.today().isoformat(),
        }

    return render(request, 'balance_sheet.html', context)


@login_required
def financials(request):
    return render(request, 'finance_page.html')

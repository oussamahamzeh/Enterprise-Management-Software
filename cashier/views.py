import json
import os
import re
import subprocess
import textwrap
import logging
import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, get_object_or_404, redirect

from InvMngSys.settings import MEDIA_ROOT
from Inventory.models import Transaction, Item, Client
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, units
from django.http import JsonResponse
from openpyxl.worksheet.page import PageMargins  # , PageSetup
# from openpyxl.worksheet.print_options import PrintPageSetup
from django.contrib import messages
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from .models import InvoiceNumber
from openpyxl.worksheet.page import PrintPageSetup
from django.http import JsonResponse

logger = logging.getLogger(__name__)


def create_transactions(request):
    print("Creating Transaction function!")
    transaction_list = []
    if request.method == 'POST':
        try:

            data = json.loads(request.body)  # Fetch the JSON data from the request
            transactions = data.get('transactions', [])  # Parse the JSON data
            filtered_transactions = [transaction for transaction in transactions if
                                     transaction['quantity'] not in ('', '0')]
            if not filtered_transactions:
                return JsonResponse({'success': True})
            # print("Received JSON data:", data)
            # print("Filtered transactions:", filtered_transactions)

            # Retrieve discount and subtotal values from the form
            discount_percentage = float(data.get('discount', 0)) / 100
            subtotal = float(data.get('subtotal', 0))
            error_messages = []
            # Now you can iterate through the transactions and perform the necessary actions
            for transaction in filtered_transactions:

                # client = Client.objects.get(name="Guest")
                type = 'transfer'
                try:
                    item = Item.objects.get(code=transaction['code'])
                except ObjectDoesNotExist:
                    return JsonResponse(
                        {'success': False, 'error': f'Item with code {transaction["code"]} does not exist'})

                quantity = transaction['quantity']
                selling_price = float(transaction['selling_price']) * (1 - discount_percentage)

                # Calculate the discount amount based on the subtotal
                discount = ((item.selling_price) - float(selling_price)) * int(quantity)

                profit = ((item.selling_price - item.purchase_cost - item.shipping_cost) * int(quantity)) - discount
                TVA = round(float(selling_price) * 0.11 * int(quantity), 2)

                transaction = Transaction(quantity=quantity, item=item, selling_price=selling_price,
                                          user=request.user.username, discount=discount, type=type, profit=profit,
                                          TVA=TVA)
                transaction.save()
                transaction_list.append(transaction)

                if item.quantity < int(quantity):
                    error_messages.append(f"Item {item.name} has insufficient quantity.")

                item.quantity = item.quantity - int(quantity)
                item.save()

            print("Transactions List:", transaction_list)

            # Create a receipt Excel
            receipt(transaction_list)

            # Clear results
            del request.session['search_results']  # Clear the session key
            request.session.modified = True  # Mark session as modified to save changes
            if error_messages:
                return JsonResponse({'success': False, 'errors': error_messages})

            return JsonResponse({'success': True})  # redirect('cashier:search_item')
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})

    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_previous_counter(directory):
    counter_pattern = re.compile(r'Receipt_(\d+)\.xlsx')
    existing_counters = []

    for filename in os.listdir(directory):
        match = counter_pattern.match(filename)
        if match:
            existing_counters.append(int(match.group(1)))

    if existing_counters:
        return max(existing_counters) + 1
    else:
        return 1


def receipt(transactions):
    if transactions:
        print("Creating Receipt function!")

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(BASE_DIR, 'media/Templates/receipt_template.xlsx')
        save_path = os.path.join(BASE_DIR, 'media/Receipts')

        # counter = get_previous_counter(save_path)
        if not os.path.exists(save_path):
            # If it doesn't exist, create the directory
            os.makedirs(save_path)

        counter = InvoiceNumber.get_next_invoice_number()

        template_workbook = openpyxl.load_workbook(template_path)
        worksheet = template_workbook[template_workbook.sheetnames[0]]

        # Apply center alignment to the cell
        alignment = Alignment(horizontal='center', vertical='center')
        font = Font(name='Arial', size=11, bold=True, italic=False)  # , color="0033CC")  # Customize font settings
        font02 = Font(name='Arial', size=12, bold=True, italic=True)  # , color="0033CC")  # Customize font settings

        # Get the current datetime
        current_datetime = datetime.now()

        # Format the datetime as a string
        formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        worksheet.cell(row=7, column=2, value=formatted_datetime)

        worksheet.cell(row=8, column=2, value=counter)
        worksheet.cell(row=8, column=2, value=counter).alignment = alignment

        # Adjust column width to fit the content

        column_letter = get_column_letter(1)  # Column A
        worksheet.column_dimensions[column_letter].width = 20  # Adjust the width as needed

        data_start_row = 11
        total_sum = 0
        added_rows = 0
        discount = 0
        rounded_discount = 0
        total_quantity = 0
        # Write the transactions to the worksheet
        for row_num, transaction in enumerate(transactions, data_start_row):
            added_rows = added_rows + 1
            worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity)
            worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity).alignment = alignment
            total_quantity += int(transaction.quantity)
            if transaction.discount < 0:
                temp_value = abs(float(transaction.item.selling_price) - float(transaction.selling_price))
                worksheet.cell(row=row_num + added_rows, column=3, value=transaction.item.selling_price + temp_value)
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=transaction.item.selling_price + temp_value).alignment = alignment
            else:
                worksheet.cell(row=row_num + added_rows, column=3, value=transaction.item.selling_price)
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=transaction.item.selling_price).alignment = alignment
            lines = line_split(transaction.item.name)
            for idx, line in enumerate(lines):
                if idx != 0:
                    added_rows = added_rows + 1
                worksheet.cell(row=row_num + added_rows, column=1, value=line)

            # worksheet.row_dimensions[row_num].auto_size = True
            total_sum = total_sum + (float(transaction.selling_price) * int(transaction.quantity))

            if transaction.discount > 0:
                discount = ((float(transaction.discount) / float(transaction.quantity)) / float(
                    transaction.item.selling_price)) * 100
                rounded_discount = round(discount, 2)
                added_rows = added_rows + 1
                worksheet.cell(row=row_num + added_rows, column=1, value="Discount " + str(rounded_discount) + " %")
                worksheet.cell(row=row_num + added_rows, column=1,
                               value="Discount " + str(rounded_discount) + " %").alignment = alignment
                worksheet.cell(row=row_num + added_rows, column=2, value="-")
                worksheet.cell(row=row_num + added_rows, column=2, value="-").alignment = alignment
                worksheet.cell(row=row_num + added_rows, column=3,
                               value=float(transaction.item.selling_price) - float(transaction.selling_price))
                worksheet.cell(row=row_num + added_rows, column=3, value=float(transaction.item.selling_price) - float(
                    transaction.selling_price)).alignment = alignment

        total_sum = round(total_sum, 2)
        final_row = len(transactions) + 10 + added_rows
        worksheet.cell(row=final_row + 2, column=1, value="Sub-total")
        worksheet.cell(row=final_row + 2, column=1, value="Sub-total").font = font
        worksheet.cell(row=final_row + 2, column=2, value=total_sum * 0.89)
        worksheet.cell(row=final_row + 2, column=3, value="USD")
        worksheet.cell(row=final_row + 3, column=1, value="TVA")
        worksheet.cell(row=final_row + 3, column=1, value="TVA").font = font
        worksheet.cell(row=final_row + 3, column=2, value=total_sum * 0.11)
        worksheet.cell(row=final_row + 3, column=3, value="USD")
        worksheet.cell(row=final_row + 5, column=1, value="Total")
        worksheet.cell(row=final_row + 5, column=1, value="Total").font = font
        worksheet.cell(row=final_row + 5, column=2, value=total_sum)
        worksheet.cell(row=final_row + 5, column=3, value="USD")
        worksheet.cell(row=final_row + 6, column=1, value="Quantity")
        worksheet.cell(row=final_row + 6, column=1, value="Quantity").font = font
        worksheet.cell(row=final_row + 6, column=2, value=total_quantity)
        if total_quantity == 1:
            worksheet.cell(row=final_row + 6, column=3, value="Item")
        else:
            worksheet.cell(row=final_row + 6, column=3, value="Items")
        worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN")
        worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN").font = font02
        worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!")
        worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!").font = font02

        # Set print page options
        worksheet.print_options.horizontalCentered = True  # Center horizontally
        worksheet.print_options.verticalCentered = False  # Optional: Center vertically
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT  # Portrait orientation

        # Set print area
        rows_to_print = final_row + 12
        print_area = 'A1:C{}'.format(rows_to_print)
        worksheet.print_area = print_area

        # Set custom page size in millimeters
        worksheet.page_setup.paperSize = 0x11  # Custom paper size code for Excel (see below)

        # Set page margins
        page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
        worksheet.page_margins = page_margins

        # Apply page setup for "Fit to One Page"
        # page_setup = worksheet.page_setup or PrintPageSetup()
        worksheet.print_options.fitToPage = True
        worksheet.print_options.fitToHeight = 1
        worksheet.print_options.fitToWidth = 1

        # Choose a directory to save the file
        file_name = f'Receipt_{counter}.xlsx'
        full_file_path = os.path.join(save_path, file_name)

        # Save the workbook to the chosen location
        template_workbook.save(full_file_path)

        print_receipt(full_file_path)

        return HttpResponse(status=204)  # Return an empty response with status code 204 (No Content)
    else:
        return HttpResponse(status=400)  # Return an empty response with status code 400 (Bad Request)


def line_split(name, max_width=24):
    # Split the item name into lines
    lines = []
    current_line = ""
    words = name.split()
    for word in words:
        if len(current_line) + len(word) + 1 <= max_width:  # +1 for space between words
            current_line += word + " "
        else:
            lines.append(current_line)
            current_line = word + " "
    lines.append(current_line)  # Add the last line

    return lines


@login_required
def cashier(request):
    # Handle item code search and item details fetching here
    return render(request, 'cashier.html')


@login_required
def wholesale(request):
    # Handle item code search and item details fetching here
    return render(request, 'wholesale.html')


def clear_results(request):
    if 'search_results' in request.session:
        del request.session['search_results']  # Clear the session key
        request.session.modified = True  # Mark session as modified to save changes
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def search_item(request):
    if request.method == 'POST':
        code_input = request.POST.get('code')
        quantity = int(request.POST.get('quantity', 1))  # Default to 1 if not provided
        items = Item.objects.filter(code=code_input)

        if 'search_results' not in request.session:
            request.session['search_results'] = []

        found_existing_item = False

        for result in request.session['search_results']:
            if result['item']['code'] == code_input:
                result['quantity'] += quantity
                found_existing_item = True
                break

        if not found_existing_item:
            for item in items:
                item_data = {
                    'code': item.code,
                    'name': item.name,
                    'selling_price': item.selling_price,
                }
                request.session['search_results'].append({'item': item_data, 'quantity': quantity})

        request.session.modified = True  # Mark session as modified to ensure changes are saved

    results = request.session.get('search_results', [])
    context = {'results': results}

    template_name = 'cashier.html'

    return render(request, template_name, context)


def update_results(request):
    logger.info("update_results view called!")
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))  # Decode the JSON data
            field_name = data.get('field')
            field_value = data.get('value')
            item_code = data.get('code')
            print(data)
            print(request.session['search_results'])
            for result in request.session.get('search_results', []):
                if result['item']['code'] == item_code:
                    # Update the quantity or selling_price based on the field name
                    if field_name == 'quantity':
                        result['quantity'] = field_value
                    elif field_name == 'price':
                        result['item']['selling_price'] = field_value

                    request.session.modified = True  # Mark session as modified to ensure changes are saved
                    print(request.session['search_results'])
                    return JsonResponse({'status': 'success'})

                # If no matching item is found, return an error
            return JsonResponse({'status': 'error', 'message': 'Item not found with code {}'.format(item_code)})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'})

    return HttpResponseBadRequest('Invalid request method')


def print_receipt(path):
    # Open the saved file using the 'start' command on Windows
    try:
        subprocess.Popen(['start', path], shell=True)  # os.startfile(path, "print")
    except Exception as e:
        print(f"Error printing file: {e}")


def wholesale_search_item(request):
    if request.method == 'POST':
        code_input = request.POST.get('code')
        quantity = int(request.POST.get('quantity'))  # Default to 1 if not provided
        items = Item.objects.filter(code=code_input)

        if 'search_results' not in request.session:
            request.session['search_results'] = []

        found_existing_item = False

        for result in request.session['search_results']:
            if result['item']['code'] == code_input:
                result['quantity'] += quantity
                found_existing_item = True
                break

        if not found_existing_item:
            for item in items:
                item_data = {
                    'code': item.code,
                    'name': item.name,
                    'selling_price': item.selling_price,
                }
                request.session['search_results'].append({'item': item_data, 'quantity': quantity})

        request.session.modified = True  # Mark session as modified to ensure changes are saved

    results = request.session.get('search_results', [])
    context = {'results': results}

    template_name = 'wholesale.html'

    return render(request, template_name, context)


def wholesale_create_transactions(request):
    print("Creating Wholesale Transaction function!")
    transaction_list = []
    if request.method == 'POST':
        try:

            data = json.loads(request.body)  # Fetch the JSON data from the request
            transactions = data.get('transactions', [])  # Parse the JSON data
            filtered_transactions = [transaction for transaction in transactions if
                                     transaction['quantity'] not in ('', '0')]
            if not filtered_transactions:
                return JsonResponse({'success': True})

            # print("Received JSON data:", data)
            # print("Filtered transactions:", filtered_transactions)

            # Retrieve discount and subtotal values from the form
            margin = float(data.get('margin', 0)) / 100
            subtotal = float(data.get('subtotal', 0))
            error_messages = []
            # Now you can iterate through the transactions and perform the necessary actions
            for transaction in filtered_transactions:

                # client = Client.objects.get(name="Guest")
                type = 'wholesale'
                try:
                    item = Item.objects.get(code=transaction['code'])
                except ObjectDoesNotExist:
                    return JsonResponse(
                        {'success': False, 'error': f'Item with code {transaction["code"]} does not exist'})

                quantity = transaction['quantity']
                selling_price = round((item.purchase_cost + item.shipping_cost) * (1 + margin), 2)
                # Calculate the discount amount based on the subtotal
                discount = ((item.selling_price) - float(selling_price)) * int(quantity)

                profit = ((item.selling_price - item.purchase_cost - item.shipping_cost) * int(quantity)) - discount
                TVA = round(float(selling_price) * 0.11 * int(quantity), 2)

                transaction = Transaction(quantity=quantity, item=item, selling_price=selling_price,
                                          user=request.user.username, discount=discount, type=type, profit=profit,
                                          TVA=TVA)
                transaction.save()
                transaction_list.append(transaction)

                if item.quantity < int(quantity):
                    error_messages.append(f"Item {item.name} has insufficient quantity.")

                item.quantity = item.quantity - int(quantity)
                item.save()

            print("Transactions List:", transaction_list)

            # Create a receipt Excel
            wholesale_receipt(transaction_list)

            # Clear results
            del request.session['search_results']  # Clear the session key
            request.session.modified = True  # Mark session as modified to save changes
            if error_messages:
                return JsonResponse({'success': False, 'errors': error_messages})

            return JsonResponse({'success': True})  # redirect('cashier:search_item')
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})

    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


def wholesale_receipt(transactions):
    print("Creating Wholesale Receipt function!")

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(BASE_DIR, 'media/Templates/wholesale_template.xlsx')
    save_path = os.path.join(BASE_DIR, 'media/Receipts')
    # counter = get_previous_counter(save_path)

    if not os.path.exists(save_path):
        # If it doesn't exist, create the directory
        os.makedirs(save_path)

    counter = InvoiceNumber.get_next_invoice_number()

    # Create a new workbook and worksheet
    template_workbook = openpyxl.load_workbook(template_path)
    worksheet = template_workbook[template_workbook.sheetnames[0]]

    # Apply center alignment to the cell
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='Arial', size=11, bold=True, italic=False)  # , color="0033CC")  # Customize font settings
    font02 = Font(name='Arial', size=12, bold=True, italic=True)  # , color="0033CC")  # Customize font settings

    # Get the current datetime
    current_datetime = datetime.now()

    # Format the datetime as a string
    formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
    worksheet.cell(row=7, column=2, value=formatted_datetime)

    worksheet.cell(row=8, column=2, value=counter)
    worksheet.cell(row=8, column=2, value=counter).alignment = alignment

    # Adjust column width to fit the content
    column_letter = get_column_letter(1)  # Column A
    worksheet.column_dimensions[column_letter].width = 20  # Adjust the width as needed

    data_start_row = 11
    total_sum = 0
    added_rows = 0
    discount = 0
    rounded_discount = 0
    total_quantity = 0
    # Write the transactions to the worksheet
    for row_num, transaction in enumerate(transactions, data_start_row):
        added_rows = added_rows + 1
        worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity)
        worksheet.cell(row=row_num + added_rows, column=2, value=transaction.quantity).alignment = alignment
        total_quantity += int(transaction.quantity)
        worksheet.cell(row=row_num + added_rows, column=3, value=transaction.selling_price)
        worksheet.cell(row=row_num + added_rows, column=3, value=transaction.selling_price).alignment = alignment
        lines = line_split(transaction.item.name)
        for idx, line in enumerate(lines):
            if idx != 0:
                added_rows = added_rows + 1
            worksheet.cell(row=row_num + added_rows, column=1, value=line)

        # worksheet.row_dimensions[row_num].auto_size = True
        total_sum = total_sum + (float(transaction.selling_price) * int(transaction.quantity))

    total_sum = round(total_sum, 2)
    final_row = len(transactions) + 10 + added_rows
    worksheet.cell(row=final_row + 2, column=1, value="Sub-total")
    worksheet.cell(row=final_row + 2, column=1, value="Sub-total").font = font
    worksheet.cell(row=final_row + 2, column=2, value=total_sum * 0.89)
    worksheet.cell(row=final_row + 2, column=3, value="USD")
    worksheet.cell(row=final_row + 3, column=1, value="TVA")
    worksheet.cell(row=final_row + 3, column=1, value="TVA").font = font
    worksheet.cell(row=final_row + 3, column=2, value=total_sum * 0.11)
    worksheet.cell(row=final_row + 3, column=3, value="USD")
    worksheet.cell(row=final_row + 5, column=1, value="Total")
    worksheet.cell(row=final_row + 5, column=1, value="Total").font = font
    worksheet.cell(row=final_row + 5, column=2, value=total_sum)
    worksheet.cell(row=final_row + 5, column=3, value="USD")
    worksheet.cell(row=final_row + 6, column=1, value="Quantity")
    worksheet.cell(row=final_row + 6, column=1, value="Quantity").font = font
    worksheet.cell(row=final_row + 6, column=2, value=total_quantity)
    if total_quantity == 1:
        worksheet.cell(row=final_row + 6, column=3, value="Item")
    else:
        worksheet.cell(row=final_row + 6, column=3, value="Items")
    worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN")
    worksheet.cell(row=final_row + 9, column=1, value="        PLEASE VISIT US AGAIN").font = font02
    worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!")
    worksheet.cell(row=final_row + 10, column=1, value="                   Thank You!").font = font02

    # Set print page options
    worksheet.print_options.horizontalCentered = True  # Center horizontally
    worksheet.print_options.verticalCentered = False  # Optional: Center vertically
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT  # Portrait orientation

    # Set print area
    rows_to_print = final_row + 12
    print_area = 'A1:C{}'.format(rows_to_print)
    worksheet.print_area = print_area

    # Set custom page size in millimeters
    worksheet.page_setup.paperSize = 0x11  # Custom paper size code for Excel (see below)

    # Set page margins
    page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    worksheet.page_margins = page_margins

    # Apply page setup for "Fit to One Page"
    worksheet.print_options.fitToPage = True
    worksheet.print_options.fitToHeight = 1
    worksheet.print_options.fitToWidth = 1

    # Choose a directory to save the file
    file_name = f'Receipt_{counter}.xlsx'
    full_file_path = os.path.join(save_path, file_name)

    # Save the workbook to the chosen location
    template_workbook.save(full_file_path)
    print_receipt(full_file_path)

    # worksheet.PrintOut()  # Print the entire workbook

    return HttpResponse(status=204)  # Return an empty response with status code 204 (No Content)
